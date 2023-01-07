from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import config
import datetime
import smtplib
import openpyxl
import win10toast

#   ACCESS TO EXCEL FILE

excel_file = openpyxl.load_workbook('mailing.xlsx')
main_list = excel_file['Main']


def main_function():

    #   FUNCTION FOR SENDING MAIL

    def send_email(subject, text):
        try:
            msg = MIMEMultipart()
            msg['From'] = config.login
            msg['To'] = email
            msg['Subject'] = subject
            msg.attach(MIMEText(text, 'plain'))

            server = smtplib.SMTP_SSL("smtp.beget.ru", 465)
            server.ehlo('logist@m-podarkov.ru')
            server.login(config.login, config.password)
            server.auth_plain()
            server.send_message(msg)
            server.quit()

        except:
            print('\nCYCLE IS END WITH PROBLEMS')

    i = 1
    r = 2

    # DATA SEARCH

    while i < main_list.max_row:

        #   FIND CELL

        track_num = main_list.cell(row=r, column=2).value
        email = main_list.cell(row=r, column=1).value
        date = main_list.cell(row=r, column=4).value
        tk = ''
        web_tk = ''

        #   PRINT TK

        if main_list.cell(row=r, column=3).value == "СДЭК":
            tk = config.name_tk[1]

        elif main_list.cell(row=r, column=3).value == "ТК КИТ":
            tk = config.name_tk[0]
        else:
            tk = config.name_tk[2]

        #     PRINT WEB TK

        if main_list.cell(row=r, column=3).value == "СДЭК":
            web_tk = config.link_web_tk[1] + str(track_num)
        elif main_list.cell(row=r, column=3).value == "ТК КИТ":
            web_tk = config.link_web_tk[0]
        else:
            web_tk = config.link_web_tk[2]

        #   LETTER FOR SENDING

        message = f"Добрый день!\n" \
                  f"Ваш заказ новогодних подарков был отправлен транспортной компанией {tk}.\n" \
                  f"Ориентировочная дата доставки {date.strftime('%d.%m.%Y')}\n" \
                  f"Трек номер вашего заказа: {track_num}\n" \
                  f"Ссылка для отслеживания {web_tk}\n" \
                  f"Спасибо за выбор нашей компании!" \
                  f"\n\n" \
                  f"--\n" \
                  f"С уважением, Гидревич Дмитрий\n" \
                  f"https://mir-novogodnih-podarkov.ru/"

        send_email("Доставка новогодних подарков", message)
        i += 1
        r += 1
        print('Completed: ' + str(i - 1))

toast = win10toast.ToastNotifier()
toast.show_toast(title='Рассылка завершена', msg='Отправлено ' + str(main_list.max_row - 1) + ' сообщений', duration=10)
print('COMPLETED ' + str(main_list.max_row - 1) + ' OPERATIONS')

# checking the number of rows
def check_num_rows ():
    print(main_list.max_row)